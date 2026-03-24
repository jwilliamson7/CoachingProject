#!/usr/bin/env python
"""Generate Excel workbook with all 12 paper tables for IJCSS submission.
Each table on its own sheet. TNR 12pt, bold headers, horizontal lines only."""

import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment

TNR = Font(name="Times New Roman", size=10)
TNR_BOLD = Font(name="Times New Roman", size=10, bold=True)
THIN = Side(style="thin")
NO_SIDE = Side(style=None)
TOP_BOTTOM = Border(top=THIN, bottom=THIN, left=NO_SIDE, right=NO_SIDE)
BOTTOM_ONLY = Border(top=NO_SIDE, bottom=THIN, left=NO_SIDE, right=NO_SIDE)
NO_BORDER = Border(top=NO_SIDE, bottom=NO_SIDE, left=NO_SIDE, right=NO_SIDE)


def add_sheet(wb, name, headers, rows, col_widths=None):
    ws = wb.create_sheet(title=name)

    # Headers (row 1): bold, top+bottom border
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = TNR_BOLD
        c.border = TOP_BOTTOM
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for i, row_data in enumerate(rows, 2):
        is_last = (i == len(rows) + 1)
        for j, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = TNR
            c.border = BOTTOM_ONLY if is_last else NO_BORDER
            c.alignment = Alignment(vertical="center")

    # Column widths
    if col_widths:
        for j, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

    return ws


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ---- TABLE 1 ----
    add_sheet(wb, "Table 1",
        ["Category", "Features", "Count", "Description"],
        [
            ["Core Experience", "1\u20138", 8, "Age, prior HC hires, years at each coaching level"],
            ["OC Statistics", "9\u201341", 33, "Team offensive performance during OC tenure"],
            ["DC Statistics", "42\u201374", 33, "Opponent offensive performance during DC tenure"],
            ["HC Team Stats", "75\u2013107", 33, "Team offensive performance during HC tenure"],
            ["HC Opponent Stats", "108\u2013140", 33, "Opponent offensive performance during HC tenure"],
            ["Hiring Team Context", "141\u2013150", 10, "Hiring team\u2019s recent performance metrics"],
        ],
        [22, 12, 8, 48])

    # ---- TABLE 2 ----
    add_sheet(wb, "Table 2",
        ["# Features", "QWK [95% CI]", "MAE [95% CI]", "Adj. Acc. [95% CI]", "Macro F1 [95% CI]"],
        [
            ["5",   "0.473 [.454, .493]", "0.524 [.509, .538]", "90.5% [89.8, 91.2]", "0.544 [.534, .554]"],
            ["10",  "0.587 [.572, .602]", "0.434 [.422, .447]", "93.5% [93.0, 93.9]", "0.605 [.595, .616]"],
            ["20",  "0.707 [.695, .719]", "0.336 [.324, .347]", "96.8% [96.4, 97.2]", "0.673 [.663, .684]"],
            ["30",  "0.729 [.715, .744]", "0.314 [.300, .328]", "97.3% [96.9, 97.7]", "0.691 [.677, .704]"],
            ["40*", "0.744 [.731, .757]", "0.307 [.294, .320]", "98.1% [97.8, 98.5]", "0.691 [.679, .703]"],
            ["50",  "0.726 [.714, .739]", "0.317 [.304, .329]", "97.4% [97.1, 97.8]", "0.688 [.675, .700]"],
            ["60",  "0.719 [.704, .734]", "0.325 [.312, .339]", "97.5% [97.1, 97.9]", "0.678 [.666, .690]"],
            ["70",  "0.722 [.709, .735]", "0.325 [.313, .337]", "97.8% [97.4, 98.1]", "0.675 [.664, .686]"],
            ["80",  "0.726 [.713, .738]", "0.325 [.313, .337]", "98.0% [97.7, 98.3]", "0.673 [.662, .684]"],
            ["90",  "0.713 [.700, .726]", "0.330 [.317, .342]", "97.5% [97.2, 97.8]", "0.674 [.663, .685]"],
            ["100", "0.707 [.693, .720]", "0.339 [.325, .352]", "97.6% [97.2, 97.9]", "0.664 [.651, .677]"],
            ["110", "0.699 [.686, .712]", "0.348 [.336, .360]", "97.6% [97.2, 98.0]", "0.653 [.643, .664]"],
            ["120", "0.696 [.683, .710]", "0.346 [.334, .358]", "97.4% [97.0, 97.7]", "0.658 [.647, .670]"],
            ["130", "0.693 [.679, .708]", "0.351 [.338, .364]", "97.3% [97.0, 97.7]", "0.652 [.641, .663]"],
            ["140", "0.688 [.674, .701]", "0.352 [.340, .364]", "97.2% [96.8, 97.5]", "0.653 [.642, .664]"],
            ["150", "0.692 [.678, .707]", "0.350 [.337, .363]", "97.3% [96.9, 97.6]", "0.656 [.643, .668]"],
        ],
        [12, 22, 22, 22, 22])

    # ---- TABLE 3 ----
    add_sheet(wb, "Table 3",
        ["Rank", "Feature", "Imp.", "95% CI", "Rank", "Feature", "Imp.", "95% CI"],
        [
            [1,  "3D% (HC)",        ".081", "[.077, .085]", 21, "4D Att (HC Opp)",   ".016", "[.015, .018]"],
            [2,  "3D Att (HC Opp)", ".073", "[.069, .077]", 22, "Int (HC)",          ".016", "[.014, .018]"],
            [3,  "RZ Att (DC)",     ".067", "[.064, .070]", 23, "RZ Att (HC Opp)",   ".015", "[.014, .017]"],
            [4,  "4D% (HC Opp)",    ".062", "[.058, .065]", 24, "Yds Off (Hire)",    ".013", "[.012, .015]"],
            [5,  "Sc% (HC Opp)",    ".054", "[.051, .057]", 25, "Pen 1D (DC)",       ".013", "[.012, .014]"],
            [6,  "Rush Att (HC Opp)",".052","[.049, .055]", 26, "Rush 1D (HC Opp)",  ".013", "[.011, .014]"],
            [7,  "3D% (HC Opp)",    ".052", "[.048, .055]", 27, "TO Forced (Hire)",  ".012", "[.011, .014]"],
            [8,  "Yds/Dr (HC)",     ".047", "[.044, .050]", 28, "Pass TD (HC)",      ".012", "[.011, .013]"],
            [9,  "Yds/Dr (HC Opp)", ".042", "[.039, .045]", 29, "TO Comm (Hire)",    ".011", "[.010, .012]"],
            [10, "Y/P (OC)",        ".030", "[.028, .032]", 30, "Yrs NFL Coor (Core)",".011","[.010, .012]"],
            [11, "Cmp (HC)",        ".030", "[.027, .032]", 31, "Pass 1D (DC)",      ".011", "[.010, .012]"],
            [12, "Pass Att (DC)",   ".029", "[.027, .032]", 32, "Pts Allow (Hire)",  ".011", "[.009, .012]"],
            [13, "TO (DC)",         ".025", "[.023, .027]", 33, "Yrs NFL Pos (Core)",".010", "[.009, .011]"],
            [14, "TO% (HC)",        ".025", "[.023, .027]", 34, "Cmp (HC Opp)",      ".010", "[.009, .011]"],
            [15, "TO (HC)",         ".021", "[.019, .023]", 35, "RZ% (HC Opp)",      ".009", "[.008, .011]"],
            [16, "Plays/Dr (OC)",   ".019", "[.017, .022]", 36, "Y/P (DC)",          ".009", "[.008, .010]"],
            [17, "#Dr (HC Opp)",    ".019", "[.017, .021]", 37, "Rush Y/A (HC Opp)", ".007", "[.007, .008]"],
            [18, "RZ Att (OC)",     ".018", "[.016, .019]", 38, "Pts (HC Opp)",      ".007", "[.006, .008]"],
            [19, "Pts (HC)",        ".017", "[.016, .019]", 39, "Y/P (Hire)",        ".007", "[.006, .008]"],
            [20, "Yds (HC)",        ".017", "[.015, .019]", 40, "Sc% (HC)",          ".006", "[.006, .007]"],
        ],
        [7, 20, 7, 14, 7, 20, 7, 14])

    # ---- TABLE 4 ----
    add_sheet(wb, "Table 4",
        ["Metric", "Mean", "95% CI"],
        [
            ["Mean Absolute Error (MAE)",      "0.307", "[0.294, 0.320]"],
            ["Quadratic Weighted Kappa (QWK)",  "0.744", "[0.731, 0.757]"],
            ["Adjacent Accuracy (\u00b11 class)","98.1%","[97.8%, 98.5%]"],
            ["Exact Accuracy",                  "71.2%", "[70.0%, 72.4%]"],
            ["Macro F1 Score",                  "0.691", "[0.679, 0.703]"],
            ["AUROC (macro OVR)",               "0.862", "[0.855, 0.869]"],
            ["Optimistic Baseline F1*",         "0.130", "\u2014"],
        ],
        [36, 10, 18])

    # ---- TABLE 5 ----
    add_sheet(wb, "Table 5",
        ["Metric", "Mean \u0394", "95% CI", "p"],
        [
            ["MAE",           "\u22120.044", "[\u22120.054, \u22120.034]", "<0.001"],
            ["QWK",           "+0.033",      "[+0.023, +0.042]",          "<0.001"],
            ["Adj. Accuracy", "+0.9pp",      "[+0.5, +1.2]pp",           "<0.001"],
            ["Exact Accuracy","+3.5pp",      "[+2.5, +4.5]pp",           "<0.001"],
            ["Macro F1",      "+0.050",      "[+0.039, +0.060]",         "<0.001"],
            ["AUROC",         "+0.022",      "[+0.017, +0.027]",         "<0.001"],
            ["Class 1 F1",    "+0.106",      "[+0.088, +0.124]",         "<0.001"],
        ],
        [16, 12, 24, 10])

    # ---- TABLE 6 ----
    add_sheet(wb, "Table 6",
        ["Category", "# Feat.", "Total |SHAP|", "95% CI", "Avg |SHAP|", "95% CI"],
        [
            ["HC Stats (Team + Opp.)", 24, "0.407", "[.400, .413]", "0.0169", "[.0167, .0172]"],
            ["DC Stats (Defense)",      6, "0.089", "[.087, .092]", "0.0149", "[.0144, .0153]"],
            ["OC Stats (Offense)",      3, "0.039", "[.036, .041]", "0.0128", "[.0121, .0136]"],
            ["Hiring Team Context",     5, "0.032", "[.030, .033]", "0.0063", "[.0060, .0066]"],
            ["Core Experience",         2, "0.012", "[.012, .013]", "0.0062", "[.0058, .0066]"],
        ],
        [24, 9, 14, 14, 12, 16])

    # ---- TABLES 7-10 (Feature descriptions) ----
    stats = ["points scored", "yards", "yards/play", "turnovers", "1st downs",
             "passing completions", "passing attempts", "passing yards", "passing touchdowns",
             "passing interceptions", "NY/A", "passing first downs",
             "rushing attempts", "rushing yards", "rushing touchdowns",
             "rush yards per play", "rushing 1st downs",
             "number of penalties", "penalty yards", "penalty 1st downs",
             "number of drives", "scoring percentage", "turnover percentage",
             "drive duration", "plays per drive", "yards per drive", "points per drive",
             "number of 3rd down attempts", "third down conversion percentage",
             "number of 4th down attempts", "4th down conversion percentage",
             "red zone attempts", "red zone percentage"]

    t7 = [[1, "Age at hiring"],
           [2, "Number of times previously hired as head coach"],
           [3, "Years\u2019 experience as college position coach"],
           [4, "Years\u2019 experience as college coordinator"],
           [5, "Years\u2019 experience as college head coach"],
           [6, "Years\u2019 experience as NFL position coach"],
           [7, "Years\u2019 experience as NFL coordinator"],
           [8, "Years\u2019 experience as NFL head coach"]]
    for i, s in enumerate(stats):
        t7.append([9 + i, f"During years as NFL OC, team\u2019s average {s}"])
    add_sheet(wb, "Table 7", ["No.", "Feature Description"], t7, [6, 56])

    t8 = [[42 + i, f"During years as NFL DC, opponent team\u2019s average {s}"] for i, s in enumerate(stats)]
    add_sheet(wb, "Table 8", ["No.", "Feature Description"], t8, [6, 56])

    t9 = [[75 + i, f"During years as NFL HC, team\u2019s average {s}"] for i, s in enumerate(stats)]
    add_sheet(wb, "Table 9", ["No.", "Feature Description"], t9, [6, 56])

    t10 = [[108 + i, f"During years as NFL HC, opponent team\u2019s average {s}"] for i, s in enumerate(stats)]
    for num, desc in [
        (141, "Hiring team\u2019s average winning percent in previous two years"),
        (142, "Hiring team\u2019s average points scored in previous two years"),
        (143, "Hiring team\u2019s average points allowed in previous two years"),
        (144, "Hiring team\u2019s average yards of offense in previous two years"),
        (145, "Hiring team\u2019s average yards of offense allowed in previous two years"),
        (146, "Hiring team\u2019s average yards/play in previous two years"),
        (147, "Hiring team\u2019s average yards/play allowed in previous two years"),
        (148, "Hiring team\u2019s average turnovers forced in previous two years"),
        (149, "Hiring team\u2019s average turnovers in previous two years"),
        (150, "Hiring team\u2019s number of playoff appearances in previous two years"),
    ]:
        t10.append([num, desc])
    add_sheet(wb, "Table 10", ["No.", "Feature Description"], t10, [6, 56])

    # ---- TABLE 11 ----
    add_sheet(wb, "Table 11",
        ["Hyperparameter", "Candidate Values"],
        [
            ["Number of Estimators", "{25, 50, 100, 200}"],
            ["Learning Rate", "{0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40}"],
            ["Max Depth", "{2, 3, 4}"],
            ["Gamma", "{0, 0.01, 0.05, 0.1}"],
            ["Lambda (L2 Reg.)", "{0, 0.01, 0.1, 0.5}"],
            ["Alpha (L1 Reg.)", "{0, 0.01, 0.1}"],
            ["Subsample", "{0.80, 0.85, 0.90, 0.95, 1.00}"],
            ["Colsample by Tree", "{0.80, 0.85, 0.90, 0.95, 1.00}"],
            ["Min Child Weight", "{1, 2, 3, 4, 5}"],
        ],
        [24, 44])

    # ---- TABLE 12 ----
    add_sheet(wb, "Table 12",
        ["Hyperparameter", "Value"],
        [
            ["Classification Method", "Frank-Hall Ordinal"],
            ["Optimization Metric", "Quadratic Weighted Kappa"],
            ["Base Classifier Objective", "binary:logistic"],
            ["Number of Binary Classifiers", 2],
            ["Number of Estimators", 200],
            ["Learning Rate", 0.25],
            ["Max Estimator Depth", 2],
            ["Gamma", 0],
            ["Lambda (L2 Regularization)", 0.1],
            ["Alpha (L1 Regularization)", 0.01],
            ["Subsample", 0.80],
            ["Colsample by Tree", 0.90],
            ["Minimum Child Weight", 3],
        ],
        [32, 24])

    out = "C:/Users/jonwi/Documents/Projects/CoachingProject/ijcss/tables.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
